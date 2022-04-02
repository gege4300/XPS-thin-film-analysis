import tkinter as tk
from tkinter.ttk import *
from PIL import ImageTk,Image
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root=tk.Tk()
root.title("XPS quantification")
#root.geometry("500 × 400")
root.configure(bg="black")

#import cross section parameters
Myfiles = []
for each_file in glob.glob("*.json"):
    Myfiles.append(each_file)

# open all the json files and load with json module
data = []
for i in range(len(Myfiles)):
    data.append(json.load(open(Myfiles[i])))

print(len(data[0]))
print(np.array(data[0]).size)
wb=Workbook()
wb=load_workbook('/home/qiankun/python/xpsquanti/cs/c1s.xlsx')
ws=wb.active
column_a=ws["A"]
column_b=ws["B"]

for cell in column_a:
    print(cell.value)

for cell in column_b:
    print(cell.value)





elelabel=tk.Label(root,fg="black",bg="white",text="1)select elements:")
elelabel.grid(row=0,column=1)

ele1label=tk.Label(root,fg="black",bg="white",text="element1")
ele1label.grid(row=1,column=0)
element1=Combobox(root)
element1['values']=("C1s","O1s","F1s","N1s","S2p")
element1.current(0)
element1.grid(row=1,column=1)

#add spacer
spacer1 = tk.Label(root, bg="black",text="")
spacer1.grid(row=1, column=2,padx=10) #padx pady for spacer

ele2label=tk.Label(root,fg="black",bg="white",text="element2")
ele2label.grid(row=1,column=3)
element2=Combobox(root)
element2['values']=("C1s","O1s","F1s","N1s","S2p")
element2.current(1)

element2.grid(row=1,column=4)

#row2
arealabel=tk.Label(root,fg="black",bg="white",text="2)enter peak area:")
arealabel.grid(row=2,column=1)

area1label=tk.Label(root,fg="black",bg="white",text="element1 area")
area1label.grid(row=3,column=0)
area1=tk.Entry(root, width=20)
area1.grid(row=3,column=1)

#add spacer
spacer2 = tk.Label(root, bg="black",text="")
spacer2.grid(row=3, column=2,padx=10) #padx pady for spacer

area2label=tk.Label(root,fg="black",bg="white",text="element2 area")
area2label.grid(row=3,column=3)
area2=tk.Entry(root, width=20)
area2.grid(row=3,column=4)

#model show
modellabel=tk.Label(root,fg="black",bg="white",text="3)select XPS model:")
modellabel.grid(row=5,column=0)

modelfig1=ImageTk.PhotoImage(Image.open("images/bulk1.png"))
modelfig2=ImageTk.PhotoImage(Image.open("images/layer1.png"))


def modelselection(imagenumber):
    global figlabel
    if modelxps.get() == "bulk":
        figlabel=tk.Label(image=modelfig1).grid(row=7,column=0) 
    elif modelxps.get() == "layer":
        figlabel=tk.Label(image=modelfig2).grid(row=7,column=0) 
       

figlabel=tk.Label(image=modelfig2).grid(row=7,column=0)     
##this combobox defines everything,... extremely important                                                                                               
modelxps=Combobox(root,values=("bulk","layer"))
modelxps.current(1)
modelxps.grid(row=6,column=0)
        
modelxps.bind("<<ComboboxSelected>>",modelselection)





#define clicked function
def clicked():
    if modelxps.get() == "bulk":
        resulttextlabel=tk.Label(root,text="the result is:").grid(row=5,column=3)
        res=str(float(area1.get())+float(area2.get()))
        resultslabel=tk.Label(root,fg="black",bg="white",text="area1+area2=: " + res).grid(row=6,column=3,padx=40,pady=40)
    elif modelxps.get() == "layer":
        resulttextlabel=tk.Label(root,text="the result is:").grid(row=5,column=3)
        res=str(float(area1.get())-float(area2.get()))
        resultslabel=tk.Label(root,fg="black",bg="white",text="area1+area2=: " + res).grid(row=6,column=3,padx=40,pady=40)
    elif modelxps.get() == "layer2":
        pass
    


#calculate button
calculate=tk.Button(root,bg="white",fg="black",text="calculate",command=clicked)
calculate.grid(row=4,column=4)

#show results







root.attributes('-type', 'dialog') 
root.mainloop()
